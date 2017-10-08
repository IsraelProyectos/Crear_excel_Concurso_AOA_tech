#!/usr/bin/env python
#coding=utf-8

import wx
from openpyxl import *
from openpyxl.styles import Color, PatternFill, Font, Border, colors, borders, Side
import pdb

class ReadAndSaveExcell():
	def __init__(self, pathFile):
			self.pathFile = pathFile
			self.columna_excel = [ ]
			self.todas_columnas = [ ]
			self.registro_excel_final = [ ]
			self.registros_excel_final =[ ]
			self.fields = ['EMAIL', 'CODIGO_1', 'CODIGO_2', 'CODIGO_3', 'CODIGO_4', 'CODIGO_5', 'NOMBRE_CODIGO_1',
									  'NOMBRE_CODIGO_2', 'NOMBRE_CODIGO_3', 'NOMBRE_CODIGO_4', 'NOMBRE_CODIGO_5', 	
								  'OBJ_AO_1', 'OBJ_AO_2', 'OBJ_AO_3', 'OBJ_AO_4', 'OBJ_AO_5', 
								  'OBJ_AOA_1', 'OBJ_AOA_2',	'OBJ_AOA_3', 'OBJ_AOA_4', 'OBJ_AOA_5',	
								  'EMAIL_CC', 'EMAIL_REMITENTE', 'EMAIL_CONTACTO', 'NOMBRE']
			self.email = ""
			self.i = 0
			self.z = 2
			self.columnaDelExcel = 1
			self.mensaje = ''
			self.colorMensaje = False

	#Método para crear el excel pasándole lista del metodo readExcel
	def writeExcel(self):
			try:
				email=''
				i=-1
				y=1
				x=13
				w=18
				t=6
				nombreImpacto=''


				#Recorriendo los registros con el mismo mail y insertando los registros en la misma fila
				for registro in self.todas_columnas:

					if email == registro[10]:
						self.registros_excel_final[i][y] = registro[0]
						self.registros_excel_final[i][x] = registro[12]
						self.registros_excel_final[i][w] = registro[17]
						self.registros_excel_final[i][t] = registro[5]
						if registro[11].title() != nombreImpacto.title():
							self.registros_excel_final[i][11] = ''
						x=x+1
						y=y+1
						w=w+1
						t=t+1
					else:
						email=registro[10]
						if registro[11] is not None:
							registro[11] = registro[11].title()
							nombreImpacto=registro[11].title()
						self.registros_excel_final.append(registro)
						i=i+1
						y=1
						x=13
						w=18
						t=6
				#pdb.set_trace()
				book = Workbook()
				hoja1 = book.active

				sig=1

				#Poniendo nombres de los campos de la cabecera en el excel
				for i in [self.fields[0], self.fields[1], self.fields[2], self.fields[3], self.fields[4], self.fields[5], self.fields[6],	
						  self.fields[7], self.fields[8], self.fields[9], self.fields[10], self.fields[11], 
						  self.fields[12], self.fields[13], self.fields[14], self.fields[15], self.fields[16],	
						  self.fields[17], self.fields[18], self.fields[19], self.fields[20], self.fields[21],
						  self.fields[22], self.fields[23], self.fields[24]]:
						  celda = hoja1.cell(row=1, column=sig).value = i
						  sig=sig+1

				#Cambiando el orden de los campos que saldrán en el excel
				for regs in self.registros_excel_final:
								regsInverse = [
								regs[10],
								regs[0],
								regs[1],
								regs[2],
								regs[3],
								regs[4],
								regs[5],
								regs[6],
								regs[7],
								regs[8],
								regs[9],
								regs[17],
								regs[18],
								regs[19],
								regs[20],
								regs[21],
								regs[12],
								regs[13],
								regs[14],
								regs[15],
								regs[16],
								regs[22],
								regs[23],
								regs[24],
								regs[11]]
								y=1
								for reg in regsInverse:
									celda = hoja1.cell(row=self.z, column=y).value = reg
									y+=1
								self.z+=1
				#pdb.set_trace()

				#Guardando el WorkBook donde seleccione el Usuario
				with wx.FileDialog(None, "Save XLSX file", wildcard="XLSX files (*.xlsx)|*.xlsx",
				   style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:
				  if fileDialog.ShowModal() == wx.ID_CANCEL:
				  	return

				  #Guardando en variable el path en donde se quiere guardar el archivo
				  pathname = fileDialog.GetPath()
				  try:
				  	with open(pathname, 'w') as file:
				  		colorFill = PatternFill(start_color='A8A8A8', end_color='A8A8A8', fill_type='solid')
						font = Font(color=colors.BLACK, italic=True, bold=True)
						border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

						for cell in hoja1["1:1"]:
							cell.font = font
							cell.fill = colorFill
							cell.border = border
				  		book.save(pathname)
				  		self.mensaje = 'El fichero se ha creado correctamente'
				  		self.colorMensaje = True
					  	
				  except IOError:
				  	wx.LogError("Cannot save current data in file '%s'." % pathname)

			except KeyError:
				self.mensaje = 'No se ha podido crear el fichero'
			except IndexError:
				self.mensaje = 'El formato de celdas del documento no es válido'
			except AttributeError:
				self.mensaje = 'El documento excel cargado no es el correcto'
			return(self.mensaje)

	#Método para leer excel del path elegido en el openFileDialog  
	def readExcel(self, e):

			try:
				#Cargando fichero desde textBox, obtenido de openFileDialog
				doc = load_workbook(self.pathFile)
				hoja = doc.worksheets[0]

				i=0

				#Leyendo filas del excel y guardándolas en una lista
				for fila in hoja.rows:
					if i != 0:
						for columna in fila:
							self.columna_excel.append(columna.value)

						#Campos agregados de cod_instalación
						self.columna_excel.insert(1,'')
						self.columna_excel.insert(2,'')
						self.columna_excel.insert(3,'')
						self.columna_excel.insert(4,'')

						#Campos agregados de concesionario
						self.columna_excel.insert(9,'')
						self.columna_excel.insert(10,'')
						self.columna_excel.insert(11,'')
						self.columna_excel.insert(12,'')

						#Campos agregados de objectivo_AOA
						self.columna_excel.insert(14,'')
						self.columna_excel.insert(15,'')
						self.columna_excel.insert(16,'')
						self.columna_excel.insert(17,'')

						#Campos agregados de objectivo_AO
						self.columna_excel.insert(6,'')
						self.columna_excel.insert(7,'')
						self.columna_excel.insert(8,'')
						self.columna_excel.insert(9,'')

						#Guardando la lista dentro de otra lista para tener las filas separadas
						self.todas_columnas.append(self.columna_excel)
						self.columna_excel = [ ]
					i+=1
				self.mensaje = self.writeExcel()
			except:
				self.mensaje = 'El excel cargado no es el correcto'
			return(self.mensaje)